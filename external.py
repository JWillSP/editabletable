from openpyxl import load_workbook
import io
from turmasdetalhes import prodt, get_professor_rm


def delete_rows(ws, row):
  for i in range(row, ws.max_row + 1):
    ws.delete_rows(row)

def get_turno(turma):
  if 'NOT' in turma:
    return 'NOTURNO'
  elif 'MAT' in turma:
    return 'MATUTINO'
  elif 'VES' in turma:
    return 'VESPERTINO'
  else:
    return 'INTEGRAL 7H'


def do_file2(df, turma='', prof='', comp='comp.: '):
  lista = [x[1].to_dict() for x in df.iterrows()]
  wb = load_workbook(filename='./statics/MEDIASTODASUNIDADES.xlsx')
  ws = wb.active
  turno = get_turno(turma)
  if prof:
    nome_rm = prodt[1]
    matrícula = nome_rm.get(prof)
    pro_nome = prodt[0].get(matrícula)
  else:
    pro_nome = get_professor_rm(turma, comp)
  ws['A3'] = turma + ' - ' + turno
  ws['B4'] = pro_nome
  ws['G4'] = comp
  start_row = 6
  num_rows = start_row + len(lista)
  for i, item in enumerate(lista):
    estudante  = item.get("estudante").split(' - ')[0]
    I = item.get("I")
    II = item.get("II")
    III = item.get("III")
    row = start_row + i
    ws.cell(row=row, column=1).value = estudante
    ws.cell(row=row, column=2).value = I
    ws.cell(row=row, column=3).value = II
    ws.cell(row=row, column=4).value = III

  delete_rows(ws=ws, row=num_rows + 1)
  buffer = io.BytesIO()
  wb.save(buffer)
  bytes_data = buffer.getvalue()
  buffer.close()
  return bytes_data


def do_file2_eja(df, turma='', prof='', comp='comp.: '):
  lista = [x[1].to_dict() for x in df.iterrows()]
  wb = load_workbook(filename='./statics/MEDIASTODASUNIDADES_EJA.xlsx')
  ws = wb.active
  turno = get_turno(turma)
  if prof:
    nome_rm = prodt[1]
    matrícula = nome_rm.get(prof)
    pro_nome = prodt[0].get(matrícula)
  else:
    pro_nome =get_professor_rm(turma, comp)
  ws['A3'] = turma + ' - ' + turno
  ws['B4'] = pro_nome
  ws['F4'] = comp
  start_row = 6
  num_rows = start_row + len(lista)
  for i, item in enumerate(lista):
    estudante  = item.get("estudante").split(' - ')[0]
    I = item.get("I")
    II = item.get("II")
    III = item.get("III")
    PERCURSO = item.get("PERCURSO")
    quanto_precisa = item.get('PRECISA NA III UND.')
    row = start_row + i
    ws.cell(row=row, column=1).value = estudante
    ws.cell(row=row, column=2).value = I
    ws.cell(row=row, column=3).value = II
    ws.cell(row=row, column=4).value = quanto_precisa
    ws.cell(row=row, column=5).value = III
    ws.cell(row=row, column=6).value = PERCURSO

  delete_rows(ws=ws, row=num_rows + 1)
  buffer = io.BytesIO()
  wb.save(buffer)
  bytes_data = buffer.getvalue()
  buffer.close()
  return bytes_data


def do_mapa(df, turma=''):
  if 'NOT' in turma:
    turno = 'NOTURNO'
  elif 'MAT' in turma:
    turno = 'MATUTINO'
  elif 'VES' in turma:
    turno = 'VESPERTINO'
  else:
    turno = 'INTEGRAL 7H'
  
  if "LCH" in turma:
    filename = './statics/MAPAPARCIAL1_2023_ITINERARIO_LING.xlsx'
  elif "MCN" in turma:
    filename = './statics/MAPAPARCIAL1_2023_ITINERARIO_NATURAIS.xlsx'
  elif "TI2" in turma:
    filename = './statics/MAPAPARCIAL1_2023_ITINERARIO_TRANSDISCIPLINAR.xlsx'
  elif "PIINT" in turma:
    filename = './statics/MAPAPARCIAL1_2023_INTEGRAL.xlsx'
  elif "TJ" in turma:
    filename = './statics/MAPA_PARCIAL_EJA1.xlsx'
  elif "TF" in turma:
    filename = './statics/MAPA_PARCIAL_EJA1.xlsx'
  elif 'NOT' in turma:
    filename = './statics/MAPAPARCIAL1_2023_REGULAR_NOTURNO.xlsx'
  else:
    filename = './statics/MAPAPARCIAL1_2023_REGULAR_DIURNO.xlsx'
  # drop row if index is not a number
  df = df.dropna(subset=['estudante'])
  lista = [x[1].to_dict() for x in df.iterrows()]
  wb = load_workbook(filename=filename)
  ws = wb.active
  # get all text from row 4 starting from column 2
  columns = [ws.cell(row=4, column=k).value for k in range(2, ws.max_column + 1)]

  turno = get_turno(turma)
  ws['A2'] = turma + ' - ' + turno
  start_row = 5
  num_rows = start_row + len(lista)
  for i, item in enumerate(lista):
    estudante  = item.get("estudante").split(' - ')[0]
    # try:
    # except AttributeError:
    #   continue
    # PERCURSO = item.get("PERCURSO")
    # quanto_precisa = item.get('PRECISA NA III UND.')
    row = start_row + i
    ws.cell(row=row, column=1).value = estudante
    for j, col in enumerate(columns):
      ws.cell(row=row, column=j+2).value = item.get(col)

  delete_rows(ws=ws, row=num_rows + 1)
  buffer = io.BytesIO()
  wb.save(buffer)
  bytes_data = buffer.getvalue()
  buffer.close()
  return bytes_data


def do_mapa_final(array, turma=''):
  if 'NOT' in turma:
    turno = 'NOTURNO'
  elif 'MAT' in turma:
    turno = 'MATUTINO'
  elif 'VES' in turma:
    turno = 'VESPERTINO'
  else:
    turno = 'INTEGRAL 7H'
  
  if "LCH" in turma:
    filename = './statics/MAPA_FINAL_2023_ITLCH.xlsx'
  elif "MCN" in turma:
    filename = './statics/MAPA_FINAL_2023_ITCNA.xlsx'
  elif "TI2" in turma:
    filename = './statics/MAPA_FINAL_2023_ITTRANS.xlsx'
  elif "PIINT" in turma:
    filename = './statics/MAPA_FINAL_2023_INTEGRAL.xlsx'
  elif "TJ" in turma:
    filename = './statics/MAPA_FINAL_2023_EJA.xlsx' 
  elif "TF" in turma:
    filename = './statics/MAPA_FINAL_2023_EJA.xlsx'
  elif 'NOT' in turma:
    filename = './statics/MAPA_FINAL_2023_REG_NOT.xlsx'
  else:
    filename = './statics/MAPA_FINAL_2023_REG.xlsx' 
  print(filename)
  df = array[0]
  other_df = array[1]
  df['estudante'] = df.index
  lista = [x[1].to_dict() for x in df.iterrows()]
  lista2 = [x[1].to_dict() for x in other_df.iterrows()]
  wb = load_workbook(filename=filename)
  ws = wb.active
  columns = [ws.cell(row=4, column=k).value for k in range(2, ws.max_column + 1)]
  turno = get_turno(turma)
  ws['A2'] = turma + ' - ' + turno
  start_row = 5
  num_rows = start_row + 3*len(lista)
  COLS = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
  for i, item in enumerate(lista):
    row = start_row + 3*i
    row2 = row + 1
    print(item.get("estudante"))
    estudante  = item.get("estudante").split(' - ')[0]
    # split estudante string for each two whitespaces ' '
    list_of_particles_of_name = estudante.split(' ')
    to_restore = []
    for l in range(0, len(list_of_particles_of_name), 3):
      to_restore.append(' '.join(list_of_particles_of_name[l:l+3]))
    for k, name_part in enumerate(to_restore):
      ws.cell(row=row+k, column=1).value = name_part
    for j, col in enumerate(columns):
      acumulate_value = item.get(col)
      ws.cell(row=row, column=j+2).value = acumulate_value
      ws[f'{COLS[j+1]}{row2}'] = lista2[i].get(col)

  delete_rows(ws=ws, row=num_rows)
  buffer = io.BytesIO()
  wb.save(buffer)
  bytes_data = buffer.getvalue()
  buffer.close()
  return bytes_data


def do_mapa_final_so_parecer(array, turma=''):
  if 'NOT' in turma:
    turno = 'NOTURNO'
  elif 'MAT' in turma:
    turno = 'MATUTINO'
  elif 'VES' in turma:
    turno = 'VESPERTINO'
  else:
    turno = 'INTEGRAL 7H'
  

    filename = './statics/MAPA_FINAL_2023_SO_PARECER.xlsx'
  print(filename)
  dfmaior = array[0]
  dfmenor = array[1]
  df = array[0]
  df['estudante'] = df.index
  lista = [x[1].to_dict() for x in df.iterrows()]
  wb = load_workbook(filename=filename)
  ws = wb.active
  turno = get_turno(turma)
  ws['A2'] = turma + ' - ' + turno
  start_row = 5
  num_rows = start_row + len(lista)
  for i, item in enumerate(lista):
    status = 'APROVADO(A)'
    row = start_row + i
    print(item.get("estudante"))
    estudante  = item.get("estudante").split(' - ')[0]
    ws.cell(row=row, column=1).value = estudante
    seriemaior = dfmaior.loc[item.get("estudante"), :]
    seriemenor = dfmenor.loc[item.get("estudante"), :]
    # compare to know if seriemaior has more different quantity of NaN values than seriemenor
    if_abandoned = not bool(seriemenor.count())
    print(seriemaior.count())
    print(seriemenor.count())

    if if_abandoned:
      status = 'DESISTENTE'
    else:
      only_recovered = seriemenor.dropna()
      if_has_3_values_lt_5 = only_recovered[only_recovered < 5].count() > 3
      if if_has_3_values_lt_5:
        status = 'REPROVADO(A)'

    ws.cell(row=row, column=3).value = status


  delete_rows(ws=ws, row=num_rows)
  buffer = io.BytesIO()
  wb.save(buffer)
  bytes_data = buffer.getvalue()
  buffer.close()
  return bytes_data